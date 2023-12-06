import openpyxl

# Criar uma planilha (salvar na variável book)
book = openpyxl.Workbook()

# Como visualizar páginas existentes
print(book.sheetnames)

# Criar uma página
book.create_sheet('Frutas')
book.create_sheet('Cores')

# Como selecionar uma página
frutas_page = book['Frutas']
cores_page = book['Cores']

# Como criar colunas e linhas
frutas_page.append(['Fruta', ' Quantidade', 'Preço'])
frutas_page.append(['Banana', ' 9', '0,99'])
frutas_page.append(['Maça', ' 7', '9,00'])
frutas_page.append(['Pera', ' 11', '1,00'])
frutas_page.append(['Uva', '3', '0,50'])
cores_page.append(['Primária', 'Secundária'])
cores_page.append(['Azul', 'Amarelo'])
cores_page.append(['Verde', 'Vermelho'])
cores_page.append(['Rosa', 'Preto'])
cores_page.append(['Branco', 'Laranja'])

# Salvar a planilha
book.save('Planilha de Compras.xlsx')

