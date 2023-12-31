import openpyxl


# Carregando o arquivo
book = openpyxl.load_workbook('Planilha de Compras.xlxs')

# Selecionando uma página
frutas_page = book['Frutas']

# Imprimir os dados de cada linha 
for rows in frutas_page.iter_rows(min_row=2,max_row=5):
    for cell in rows:
        print(cell.value)

# OU

# Imprimir os dados tudo na mesma linha
for rows in frutas_page.iter_rows(min_row=2, max_row=5):
    print(rows[0].value,rows[1].value,rows[2].value)